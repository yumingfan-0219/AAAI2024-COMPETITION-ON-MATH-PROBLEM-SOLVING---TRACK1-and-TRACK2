import os
import json
import datetime
import re
from transformers import AutoModelForCausalLM, AutoTokenizer, GenerationConfig
from sentence_transformers import SentenceTransformer, models
import openpyxl
from openpyxl import Workbook
import torch
import ast

# ----- Global Variables -----

current_time = datetime.datetime.now()
timestamp = current_time.strftime("%m-%d-%H-%M")  # 格式化时间戳
model_names = [
    # 'QWEN-AFTER-PRETRAIN',
    '1024-16-4nodes-NewQWen-1024-full-3epoch',
    # '1012-16-5nodes-NewQWen-1011-full-3epoch',
    # "1019-10-5nodes-NewQWen-1018-org-full-3epoch"
    # 'NEW-QWEN'
    # 'FT-GZ-0925'
    ]

finetuned_model_path = '/work/share/embed/model/output_0918_plus'
word_embedding_model = models.Transformer(finetuned_model_path, max_seq_length=512, do_lower_case=True)
pooling_model = models.Pooling(word_embedding_model.get_word_embedding_dimension(), pooling_mode='cls')
similarity_model = SentenceTransformer(modules=[word_embedding_model, pooling_model])

# 一级标签or二级标签
# content_key = "topk_chunks"
content_key = "topk_contents"

# ----- Utility Functions -----

def compute_similarity(text1, text2):
    """Compute cosine similarity between two texts."""
    embeddings_1 = similarity_model.encode(text1, normalize_embeddings=True)
    embeddings_2 = similarity_model.encode(text2, normalize_embeddings=True)
    score = embeddings_1 @ embeddings_2.T
    return score


def process_string(input_str):
    """Clean and preprocess the input string."""
    if isinstance(input_str, list):
        input_str = list(set(input_str))
        input_str_all = ''.join([f'{text}' for text in input_str])
        input_str = input_str_all
    while True:
        prev_str = input_str
        input_str = re.sub(r'\r\n|\n |\u3000| \n|\n\n|　|\xa0|<add>|<chunk_splitter>|_x000D_\n', ' ', input_str)
        input_str = re.sub(r' {2,}', ' ', input_str)
        input_str = re.sub(r'\n{2,}', '\n', input_str)
        # input_str = input_str.strip()
        if prev_str == input_str:
            break
    return input_str

def format_content(entry):
    # 将文本列表转换为指定格式的字符串
    if isinstance(entry[content_key], list):
            content_list = entry[content_key]
    else:
        try:
            # 尝试将 content 字段的值从字符串表示的列表转换为实际的列表
            content_list = ast.literal_eval(entry[content_key])
        except SyntaxError:
            print(f"Failed to parse content field for entry: {entry}")
            content = process_string(entry[content_key])
            return f'\n<content>{content}<content>\n'
    content_list = list(set(content_list))
    formatted_content = ''.join([f'\n<content>{process_string(text)}<content>' for text in content_list])
    return formatted_content

# ----- Model Initialization -----

def init_model(model_name):
    """Initialize the model and tokenizer."""
    # model_path = "/work/share/fym/model/" + model_name
    # model_path = "/work/cache/model/NewQWen/" + model_name
    # model_path = "/work/cache/model/v0925add/" + model_name
    # model_path = '/work/share/public/weights/Qwen-7B-0925-modelscope/chat/qwen/Qwen-7B-Chat'
    model_path = "/work/cache/model/NewQWen/1024-16-4nodes-NewQWen-1024-full-3epoch"
    # model_path = "/work/share/cmj/gungzhou/model/0926-QWen-canton-chat-full-3e-5-3epoch-925-v2"
    # model_path = "/work/share/cmj/gungzhou/model/1024-QWen-chat-925-full-3e-5-3epoch-925v2data"
    model = AutoModelForCausalLM.from_pretrained(
        model_path,
        torch_dtype=torch.float16,
        device_map="auto",
        trust_remote_code=True
    )
    model.generation_config = GenerationConfig.from_pretrained(model_path)
    tokenizer = AutoTokenizer.from_pretrained(
        model_path,
        use_fast=False,
        trust_remote_code=True
    )
    return model, tokenizer


# ----- Main Processing Functions -----

def add_prompt(entry, model, tokenizer):
    """Generate response for a given entry."""
    # prompt = (
    #     "你是一个专门用于政务咨询的智能助手。"
    #     "你的任务是根据已有的政策原文和官方文件，对用户问题进行回答。请你1.仔细思考用户的真实意图，2.在政策文档中准确寻找到关键段落，3.根据段落判断用户是否符合要求，根据用户意图回答用户问题。回答需要清晰、有逻辑、有依据、亲切。"
    #     "已知政策原文和参考文档："
    #     + process_string(entry['content'])
    #     + "\n用户提问："
    #     + entry['question']
    #     + "\n请用json回答问题，格式如下：{\"intent\":给出用户的真实提问意图,\"answer\":模型的回答,\"policy\":最相关的政策原文段落，不要太长。请逐步思考。}"
    # )
    # prompt = (
    #     "你是一个专门用于政务咨询的智能助手。"
    #     "你的任务是根据已有的政策原文和官方文件，通过思维链路的格式，以亲切和友好的方式回答用户的政务问题。在提供解答时，请首先解释政策原文的相关内容，然后根据用户的具体情况，推导出针对性的回答。"
    #     "已知政策原文和参考文档："
    #     + process_string(entry['content'])
    #     + "\n用户问题："
    #     + entry['question']
    #     + "\n请回答用户提问，并根据政策原文给出相应的推理过程。你只能返回一个可解析的json格式的数据，key分别是\"policy\"，\"answer\",对应的value为参考的政策原文和模型的回复内容，对于输出结果请检查json格式的准确性，并修正结果。"
    # )
    # prompt = (
    #     "你是一个专门用于政务咨询的智能助手。"
    #     "你的任务是根据已有的政策原文和官方文件，提供一模一样的政策段落来回答用户的政务问题。你不得对政策进行解释、评判或推测。请只提供政策原文中的相关段落。"
    #     "已知政策原文和参考文档："
    #     + process_string(entry[content_key])
    #     + "\n用户问题："
    #     + entry['query']
    #     + "你不得对政策进行解释、评判或推测。请只提供政策原文中的相关段落。"
    # )
    content = format_content(entry)

    prompt = (
        "你是一个专门用于政务咨询的智能助手。"
        "你的任务是根据已有的政策原文和官方文件，提供一模一样的政策段落来回答用户的政务问题。你不得对政策进行解释、评判或推测。请只提供政策原文中的相关段落。"
        "已知政策原文和参考文档："
        + content
        + "\n用户问题：\n<question>"
        + entry['query']
        + "<question>\n请回答。"
    )
    response, _ = model.chat(tokenizer, prompt, history=None)

    print(f"Input: {prompt}")
    print(f"Output: {response}")
    print("-" * 50)  # 添加分隔线以区分每次推理
    if torch.backends.mps.is_available():
        torch.mps.empty_cache()

    similarity_score = compute_similarity(entry['answer'], response)
    is_right = 1 if similarity_score > 0.85 else 0
    return prompt, response, similarity_score, is_right


def process_file(file_path, filename, model, tokenizer, model_name):
    """Process a given file and return the results."""
    data = [('原始提问', '模型输入', '原始回答', '模型回答', '得分', '是否正确')]
    with open(os.path.join(file_path, filename), 'r', encoding='utf-8') as f:
        content = json.load(f)
        for entry in content:
            prompt, response, score, is_right = add_prompt(entry, model, tokenizer)
            data.append((entry["query"], prompt, entry["answer"], response, score, is_right))
    return data


def save_to_excel(data, sheet_name, output_file):
    """Save the results to an Excel file."""
    if os.path.exists(output_file):
        wb = openpyxl.load_workbook(output_file)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
    ws = wb.create_sheet(title=sheet_name)
    for row in data:
        ws.append(row)
    wb.save(output_file)


# ----- Main Execution -----

def main():
    """Main execution function."""
    file_path = '/work/home/project/LLaMA-Efficient-Tuning/eval_data/'
    for model_name in model_names:
        model, tokenizer = init_model(model_name)
        output_path = os.path.join('/work/home/out/跨文档检索/', model_name)
        os.makedirs(output_path, exist_ok=True)
        file = 'search_广州40条测试用例1023.json'
        # file = '测试10条.json'
        data = process_file(file_path, file, model, tokenizer, model_name)
        num_data = str(len(data))
        output_file = os.path.join(output_path, f'{timestamp}_{model_name}_{num_data}_{content_key}.xlsx')
        
        save_to_excel(data, sheet_name=file.replace('.json', ''), output_file=output_file)

        # Apply coloring (optional)
        from match_highlighter import CellColor, ExcelSheet
        sheet = ExcelSheet(output_file)
        sheet.compare('原始回答').with_column('模型回答', CellColor.purple, 10)
        output_file_color = os.path.join(output_path, f'{timestamp}_{model_name}_{num_data}_{content_key}_color.xlsx')
        sheet.export(output_file_color)
        print(f"Processed {file} for model {model_name}. Results saved to {output_file_color}")


if __name__ == "__main__":
    main()

